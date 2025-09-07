from openpyxl import load_workbook
from domain.student_record import Record

# openpyxlはExcelファイル用の専用ライブラリ。cell.valueなんてデフォルトであるわけないじゃん


def read_records(file_path: str) -> list[Record]:
    try:
        wb = load_workbook(filename=file_path)
        # inputシートを指定。
        ws = wb["input"]
    except FileNotFoundError:
        print("ファイルが見つかりません")
        return[]
    except KeyError:
        print("inputシートが見つかりません")
        return[]
    
    records = []
    for row in ws.iter_rows(min_row=2):  # ヘッダーを飛ばす
        grade = row[0].value
        student_id = row[2].value
        good_points = [cell.value for cell in row[3:7] if cell.value]
        if grade and student_id and good_points:
            # コメントはデフォでNoneが入る。
            records.append(Record(grade=grade, student_id=student_id, good_points=good_points))

    return records

def write_comments(records, filename):
    wb = load_workbook(filename)
    ws = wb["input"]

    for i, record in enumerate(records, start=2):  # 2行目から始まる前提
        ws.cell(row=i, column=8).value = record.comment  # H列 = 8列目に書くよ！

    wb.save(filename)